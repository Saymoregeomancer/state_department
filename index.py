from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from typing import List, Any
from pathlib import Path
import os
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
import random
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


from utils import (
    read_excel_as_matrix,
    write_matrix_to_excel,
    is_proxy_working,
    read_text_file_as_list,
    list_excel_files_in_dir,
    get_dir_path,
    proxy_for_selenium,
    format_for_excel
)

USER_AGENTS = [
    # === Windows ===
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",

    # === macOS ===
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",

    # === Linux ===
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
]
from datetime import datetime

def highlight_if_ukraine(file_path: str, sheet_name: str):
    wb = load_workbook(file_path)
    
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Лист '{sheet_name}' не знайдено")

    ws = wb[sheet_name]

    green_fill = PatternFill(
        start_color="90EE90",
        end_color="90EE90",
        fill_type="solid"
    )

    for row in ws.iter_rows(min_row=1):
        cell_3 = row[2]  # 3-тя комірка
        cell_2 = row[1]  # 2-га комірка

        if cell_3.value and "Ukraine" in str(cell_3.value):
            cell_2.fill = green_fill

    wb.save(file_path)




def log(text: str):
    now = datetime.now().strftime("%H:%M:%S")
    print(f"[{now}] {text}")



def check_drive(driver):
    try:
        # WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, "thead.ecl-table__head")))
        WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#overlayPanel")))
        return True    
    except:
        js_code = """
let status = document.querySelector('body')?.innerText 
let check = document.querySelector('#overlayPanel')?.innerText
if (status === 'Too Many Requests' || check === 'Too Many Requests' ) {
    // console.log('too many')
    return('too many')
}
else {
    // console.log('good')
    return('good')
}
                """
        result = driver.execute_script(js_code)
        if result == 'too many':
            log("Сайт брикається ⚠️⚠️⚠️")
            return False
        elif result == 'good':
            return True
        return False


# overlayPanel
import random

def create_chrome(proxy: str = None):
    chrome_options = Options()

    # === User-Agent rotation ===
    user_agent = random.choice(USER_AGENTS)
    chrome_options.add_argument(f"--user-agent={user_agent}")
    chrome_options.add_argument("--headless")  # якщо потрібно без GUI

    # === базові опції ===
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")

    # === proxy ===
    if proxy:
        chrome_options.add_argument(f"--proxy-server={proxy}")

    # === anti-webdriver flags ===
    chrome_options.add_experimental_option(
        "excludeSwitches", ["enable-automation"]
    )
    chrome_options.add_experimental_option(
        "useAutomationExtension", False
    )

    service = Service(chromedriver_path)
    driver = webdriver.Chrome(service=service, options=chrome_options)

    # === JS anti-webdriver ===
    driver.execute_script("""
        Object.defineProperty(navigator, 'webdriver', {
            get: () => undefined
        });
    """)

    return driver
def restart_browser(driver, proxy=None):

    try:
        driver.quit()
    except:
        pass

    time.sleep(random.uniform(2, 4))
    return create_chrome(proxy=proxy)

# ===================== Основний скрипт =====================
# print("Starting script...")
log("Початок обробки...ℹ️")

main_path = get_dir_path()
lists_list = fr'{main_path}\lists.txt'

lists_arr = read_text_file_as_list(lists_list)

proxy_file_path = fr'{main_path}\proxi.txt'
proxy_list = read_text_file_as_list(proxy_file_path)

chromedriver_path = fr"{main_path}\chromedriver.exe"


main_exel_path = list_excel_files_in_dir(main_path)[0]

init_proxy = proxy_list[random.randint(0, len(proxy_list) - 1)]
while is_proxy_working(init_proxy,3) == False:
    log(f"Перебираю проксі {init_proxy}")
    proxy_list.remove(init_proxy)
    init_proxy = proxy_list[random.randint(0, len(proxy_list) - 1)]



driver = create_chrome(proxy_for_selenium(init_proxy))


circle_count = 1

for list_item in lists_arr:
    log(f"Зараз будемо обробляти лист: {list_item}")
    data = read_excel_as_matrix(main_exel_path, list_item)
    new_data = []
    
    for i, row in enumerate(data):
            # continue  # пропускаємо порожні MRN
        if circle_count % 15 == 0:
            log(f"Трішки чекаємо, поки сервіс ображаєтсья...")
            new_proxy = proxy_list[random.randint(0, len(proxy_list) - 1)]
            while is_proxy_working(new_proxy,3) == False:
                proxy_list.remove(new_proxy)
                new_proxy = proxy_list[random.randint(0, len(proxy_list) - 1)]
            driver = restart_browser(driver, proxy_for_selenium(new_proxy))
            circle_count = 1  


        if i % 3 == 0:  # кожен 3-й рядок, починаючи з першого
            mrn = row[1]  # друга колонка
            if mrn == None:
                continue

            # if not mrn or mrn.strip() == "":
            #     new_data.append(row)
            # print(mrn)
            url = f"https://ec.europa.eu/taxation_customs/dds2/mrn/mrn_home.jsp?Lang=en&Expand=true&MRN={mrn}"
            driver.get(url)
            resp = check_drive(driver)
            while not resp:
                new_proxy = proxy_list[random.randint(0, len(proxy_list) - 1)]
                while is_proxy_working(new_proxy,3) == False:
                    log(f"Перевіряю з'єднання ⏳⏳⏳")
                    proxy_list.remove(new_proxy)
                    new_proxy = proxy_list[random.randint(0, len(proxy_list) - 1)]
                driver = restart_browser(driver, proxy_for_selenium(new_proxy))
                driver.get(url)
                resp = check_drive(driver)
            try:
                time.sleep(2)
                # Чекаємо до 15 секунд, поки таблиця завантажиться
                WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#overlayPanel")))

                js_code = """
                let data = document.querySelectorAll('tr.ecl-table__row');
                let arr = [];
                data?.forEach(el => {
                    let elem = el.querySelectorAll('.ecl-table__cell')[1]?.innerText;
                    arr.push(elem);
                });
                let newArr = arr.filter(value => typeof value === 'string')
                
                if (newArr.length > 0) {
                    // console.log(newArr.join('|'))
                    return newArr.join('|')
                }
                else if ( document.querySelector('div.form-content')) {
                    // console.log('empy')
                    return('empty')
                }
                else {
                    // console.log('error')
                    return('error')
                }



                """
                result = driver.execute_script(js_code)
            except:
                result = 'error'  # якщо таблиця не завантажилась
            # return result

            log(f"Результат: {mrn} ✅")

            circle_count += 1
            row.append(result)
            new_data.append(row)
        
        else:
            if (i+1) % 3 == 0 and row[1] != None:
                # print(type (row[1]))
                row[1] = format_for_excel(row[1])
            new_data.append(row)

    write_matrix_to_excel(main_exel_path, f"{list_item}_gotovo", new_data)
    log(f"Лист {list_item} оброблено ✅")


log('Залишилось зафарбувати комірки... 🎨')

for list_item in lists_arr:
    highlight_if_ukraine(main_exel_path, f"{list_item}_gotovo")
    log(f"Лист {list_item}_gotovo зафарбовано ✅")

log("На цьому все✅")