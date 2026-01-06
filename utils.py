from openpyxl import load_workbook, Workbook
from typing import List, Any
from pathlib import Path
import os
import sys
from pathlib import Path
from typing import List
import os
import requests


# ===================== Читання Excel =====================
def read_excel_as_matrix(file_path: str, sheet_name: str) -> List[List[Any]]:
    wb = load_workbook(filename=file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Лист '{sheet_name}' не знайдено")
    sheet = wb[sheet_name]
    return [list(row) for row in sheet.iter_rows(values_only=True)]



# ===================== Запис Excel =====================
def write_matrix_to_excel(file_path: str, sheet_name: str, data: List[List[Any]]) -> None:
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        wb = Workbook()

    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet.delete_rows(1, sheet.max_row)
    else:
        sheet = wb.create_sheet(title=sheet_name)

    for row in data:
        sheet.append(row)

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    wb.save(file_path)



# ===================== Список Excel та TXT файлів =====================
def list_excel_files_in_dir(directory: str) -> List[str]:
    dir_path = Path(directory)
    if not dir_path.is_dir():
        raise ValueError(f"{directory} is not a valid directory")
    
    return [str(dir_path / f) for f in os.listdir(dir_path) if f.lower().endswith((".xlsx", ".xls"))]






# ===================== Читання текстового файлу =====================
def read_text_file_as_list(file_path: str) -> List[str]:
    with open(file_path, "r", encoding="cp1251") as f:
        return [line.strip() for line in f]



def get_dir_path():
    
    script_path = os.path.abspath(__file__)

    if getattr(sys, 'frozen', False):
        # Якщо скрипт запущено як виконуваний файл
         script_dir = os.path.dirname(sys.executable)
    else:
        # Якщо скрипт запущено з інтерпретатора Python
         script_dir = os.path.dirname(os.path.abspath(__file__))
    return script_dir



def is_proxy_working(proxy: str, timeout: int = 7) -> bool:
    """
    Перевіряє, чи працює проксі.
    proxy: 'http://user:pass@ip:port'
    """

    proxies = {
        "http": proxy,
        "https": proxy,
    }

    try:
        response = requests.get(
            "https://httpbin.org/ip",
            proxies=proxies,
            timeout=timeout,
        )

        return response.status_code == 200

    except requests.RequestException:
        return False
    

def proxy_for_selenium(proxy_str: str) -> str:
    """
    Повертає рядок для chrome_options.add_argument()
    """
    # http://ip:port
    parts = proxy_str.split("@")[-1]
    return f"--proxy-server=http://{parts}"





from datetime import datetime

def format_for_excel(date_str):
    """
    Приймає рядок у форматі 'YYYY-MM-DD HH:MM:SS' або datetime
    Повертає рядок 'YYYY.MM.DD HH:MM:SS'
    """
    # Якщо вже datetime
    if isinstance(date_str, datetime):
        dt = date_str
    else:
        # Перетворюємо рядок у datetime
        dt = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
    
    # Повертаємо у форматі з крапками
    return dt.strftime("%Y.%m.%d %H:%M:%S")